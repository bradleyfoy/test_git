from django.shortcuts import render
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

# Create your views here.
def results(request):
    print("this is a test!")
    path = "C:/Users/foy/Desktop/test.xlsx"
    wb = load_workbook(path)
    sheet_ranges = wb['Sheet1']
    my_sheet_data = SheetData(sheet_ranges['A2'].value, sheet_ranges['B2'].value, sheet_ranges['E2'].value, )
    sheet_ranges['A10'].fill = PatternFill(start_color='FFFF0000',end_color='FFFF0000', fill_type='solid')
    wb.save("C:/Users/foy/Desktop/test.xlsx")
    my_list = []

    for row in sheet_ranges.iter_rows():
        my_sheet_data = SheetData(row[0].value, row[1].value, row[4].value )
        my_list.append(my_sheet_data)
    return render(request, 'results.html',{'my_list':my_list})

class SheetData():
    def __init__(self, name, manager, user_type):
        self.name= name
        self.manager = manager
        self.user_type = user_type

